import csv,os,sys

from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QListWidget, QLineEdit, QLabel, QDialog, \
    QPushButton, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem, QHBoxLayout,QFileDialog,QMenuBar,QMenu,QMainWindow
from PyQt6.QtGui import QPixmap, QImageReader
from PyQt6.QtCore import Qt,QSettings

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from json import dump, load
from openpyxl import load_workbook

class ElectricSymbolViewer(QWidget):
    FILE_SELECTORS = "XML Excel 2016 (*.xlsx);;Comma-separated values (*.csv);;JSON web format (*.json)"
    def __init__(self, symbols,parent:QMainWindow):
        super().__init__()

        self.setMinimumSize(720,470)

        self.symbols = symbols

        self.init_ui()


    def init_ui(self):
        layout = QVBoxLayout()

        self.title = QLabel("Electric Symbol Explorer")
        self.title.setStyleSheet("font-size:28px;font-weight:bold")

        self.utils_list = QHBoxLayout()
        self.search_input = QLineEdit(self)
        self.search_input.setPlaceholderText("Search...")
        self.search_input.textChanged.connect(self.update_list)
        self.export_button = QPushButton("Save as")
        self.export_button.clicked.connect(self.export_symbols)
        self.clear_scheme_button = QPushButton("Clear scheme")
        self.clear_scheme_button.clicked.connect(self.clear_scheme)
        self.load_button = QPushButton("Load")
        self.load_button.clicked.connect(self.load_symbols)

        self.symbol_list = QListWidget(self)
        self.update_list()

        self.symbol_description = QLabel(self)
        self.symbol_description.setText("No symbol selected")
        self.symbol_description.setWordWrap(True)

        self.action_buttons = QHBoxLayout()

        self.add_symbol_button = QPushButton("Add Symbol")
        self.add_symbol_button.clicked.connect(self.add_symbol)
        self.remove_symbol_button = QPushButton("Remove Symbol")
        self.remove_symbol_button.clicked.connect(self.remove_symbol)
        self.action_buttons.addWidget(self.add_symbol_button)
        self.action_buttons.addWidget(self.remove_symbol_button)

        layout.addWidget(self.title)
        self.utils_list.addWidget(self.search_input)
        self.utils_list.addWidget(self.export_button)
        self.utils_list.addWidget(self.load_button)
        self.utils_list.addWidget(self.clear_scheme_button)
        layout.addLayout(self.utils_list)
        layout.addWidget(self.symbol_list)
        layout.addWidget(self.symbol_description)
        layout.addLayout(self.action_buttons)

        self.symbol_list.currentItemChanged.connect(self.display_symbol_info)

        self.setLayout(layout)
        self.setWindowTitle("Electric Symbol Explorer")

    def update_list(self):
        filter_text = self.search_input.text().lower()
        self.symbol_list.clear()

        added=set()

        # First, add items that start with the filter_text
        for symbol, data in sorted(self.symbols.items()):
            if symbol.lower().startswith(filter_text) and symbol not in added:
                self.add_item(symbol,data)
                added.add(symbol)

        # Then, add items that contain the filter_text
        for symbol, data in sorted(self.symbols.items()):
            if filter_text in data["text"].lower() and symbol not in added:
                self.add_item(symbol, data)
                added.add(symbol)

    def display_symbol_info(self, current, previous):
        if current is not None:
            symbol = current.text().split(":")[0]
            description = self.symbols[symbol]["text"]
            self.symbol_description.setText(description)
    def add_item(self,symbol,data):
        self.symbol_list.addItem(f"{symbol}:{data.get('text','')}")
    def add_items(self, symbols,datas):
        for symbol in symbols:
            self.add_item(symbol, datas.get('text',"nessunad descrizione"))
    def remove_symbol(self):
        selected_item = self.symbol_list.currentItem()
        if selected_item is not None:
            symbol = selected_item.text().split(":")[0]
            del self.symbols[symbol]
            self.update_list()
    def add_symbol(self):
        dialog = AddSymbolDialog(self.symbols,self)
        dialog.exec()
    def clear_scheme(self):
        self.symbols = {}
        self.update_list()
    def export_symbols(self):
        # ask the user for a path to save the file (json or CSV)
        path, _ = QFileDialog.getSaveFileName(
            self, "Save File", "", self.FILE_SELECTORS)
        if path:
            # save the symbols to the file
            with open(path, "w", newline='') as file:
                if path.endswith(".json"):
                    dump(self.symbols, file)
                elif path.endswith(".csv"):
                    csv_writer = csv.writer(file, delimiter=',')
                    csv_writer.writerow(['Simbolo', 'Valore'])
                    for symbol, data in self.symbols.items():
                        csv_writer.writerow([symbol, data['text']])
                elif path.endswith(".xlsx"):
                    workbook = Workbook()
                    worksheet: Worksheet = workbook.active # type: ignore
                    worksheet.append(['Simbolo', 'Valore'])
                    for symbol, data in self.symbols.items():
                        worksheet.append([symbol, data['text']])
                    workbook.save(path)

    def load_symbols(self):
        # ask the user for a path to load the file, either CSV or json
        path, _ = QFileDialog.getOpenFileName(
            self, "Load File", "", self.FILE_SELECTORS)
        if path:
            # load the symbols from the file
            with open(path, "r") as file:
                if path.endswith(".json"):
                    self.symbols = load(file)
                elif path.endswith(".csv"):
                    self.symbols = {}
                    csv_reader = csv.reader(file, delimiter=',')
                    is_header_skipped = False
                    for row in csv_reader:
                        if not is_header_skipped:
                            is_header_skipped = True
                            continue
                        symbol, description = row
                        self.symbols[symbol] = {"text": description}
                elif path.endswith(".xlsx"):
                    workbook = load_workbook(path)
                    worksheet: Worksheet = workbook.active #type: ignore
                    self.symbols = {}
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        symbol, description = row
                        self.symbols[symbol] = {"text": description}
                    workbook.close()
            # update the list of symbols
            self.update_list()
            self.display_symbol_info(self.symbol_list.currentItem(), None)

class AddSymbolDialog(QDialog):
    def __init__(self, symbols,parent:ElectricSymbolViewer):
        super().__init__(parent=parent)
        self.prnt=parent;

        self.symbols = symbols

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Add Symbol")

        self.symbol_input = QLineEdit(self)
        self.symbol_input.setPlaceholderText("Symbol")

        self.description_input = QLineEdit(self)
        self.description_input.setPlaceholderText("Description")

        self.add_button = QPushButton("Add", self)
        self.add_button.clicked.connect(self.accept)

        layout = QVBoxLayout()
        layout.addWidget(self.symbol_input)
        layout.addWidget(self.description_input)
        layout.addWidget(self.add_button)

        self.setLayout(layout)
    def get_symbol_and_description(self):
        return self.symbol_input.text(), self.description_input.text()
    def accept(self):
        symbol = self.symbol_input.text()
        description = self.description_input.text()
        if symbol and description:
            self.symbols[symbol] = {"text": description}
            self.prnt.update_list()
            self.close()
            super().accept()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.settings = QSettings("Electric Symbol Explorer", "Electric Symbol Explorer")
        # get x,y, width height
        self.x_v = self.settings.value("x", 0)
        self.y_v = self.settings.value("y", 0)
        self.width_v = self.settings.value("width", 720)
        self.height_v = self.settings.value("height", 480)
        self.move(self.x_v, self.y_v)
        self.resize(self.width_v, self.height_v)

        self.init_ui()

    # save settings when window closes
    def closeEvent(self, event):
        self.settings.setValue("x", self.x())
        self.settings.setValue("y", self.y())
        self.settings.setValue("width", self.width())
        self.settings.setValue("height", self.height())
        super().closeEvent(event)
    def init_ui(self):
        self.setWindowTitle("Electric Symbol Explorer")


        self.electric_symbol_viewer = ElectricSymbolViewer({},self)

        self.menu_bar = QMenuBar(self)
        self.file_menu = QMenu("File", self)
        self.new_window = self.file_menu.addAction("New Window")
        self.new_window.triggered.connect(self.new_window_action)
        self.load_action = self.file_menu.addAction(
            "Load")
        self.load_action.setShortcut(
            "Ctrl+L")
        self.load_action.triggered.connect(self.electric_symbol_viewer.load_symbols)
        self.save_action = self.file_menu.addAction(
            "Save")
        self.save_action.setShortcut("Ctrl+S")
        self.save_action.triggered.connect(
            self.electric_symbol_viewer.export_symbols)
        self.menu_bar.addMenu(self.file_menu)
        self.setMenuBar(self.menu_bar)

        self.setCentralWidget(self.electric_symbol_viewer)
    def new_window_action(self):
        self.new_window = MainWindow()
        self.new_window.move(self.x()+24, self.y()+24)
        self.new_window.show()

if __name__ == '__main__':
    app = QApplication(sys.argv)

    electric_symbol_viewer = MainWindow()
    electric_symbol_viewer.show()

    app.exec()
